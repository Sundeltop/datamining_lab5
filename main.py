from random import random
from time import time
from typing import List

from openpyxl import load_workbook
from itertools import combinations
import matplotlib


# workbook = load_workbook('test.xlsx')
# sheet = workbook.active
# level_of_support = 3
# sets = dict()
# combo = dict()

# for rowno, rowval in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
#     invoice_number = rowval[0].value
#     if str(invoice_number).startswith("C"):
#         continue
#     stock_code = rowval[1].value
#     customer_id = rowval[6].value
#     if customer_id not in sets:
#         sets[customer_id] = str(stock_code)
#     else:
#         sets[customer_id] += ","+str(stock_code)
#
# for key in sets:
#     for item in sets[key].split(","):
#         if item not in combo:
#             combo[item] = 1
#         else:
#             combo[item] += 1
#
# above_level_list = list()
# for key in combo:
#     if combo[key] >= level_of_support:
#         above_level_list.append(key)
#
# n = 3
# for i in range(1,n):
#     if i > 1:
#         above_level_list = list(set([item for t in above_level_list for item in t]))
#
#     pairs = [",".join(map(str, comb)) for comb in combinations(above_level_list, i+1)]
#     itemset = list()
#     for pair in pairs:
#         itemset.append(pair.split(","))
#
#     combo.clear()
#     above_level_list.clear()
#     for key in sets:
#         for item in itemset:
#             if all(elem in sets[key].split(",") for elem in item):
#                 if tuple(item) not in combo:
#                     combo[tuple(item)] = 1
#                 else:
#                     combo[tuple(item)] += 1
#     for key in combo:
#         if combo[key] >= level_of_support:
#             above_level_list.append(key)
#
# print(above_level_list)


# sets[100] = "a,b,c"
# sets[200] = "b,d"
# sets[300] = "b,a,d,c"
# sets[400] = "e,d"
# sets[500] = "a,b,c,d"
# sets[600] = "f"
#
# print("initial data set: ")
# for item in sets:
#     print(sets[item])
#
# for key in sets:
#     for item in sets[key].split(","):
#         if item not in combo:
#             combo[item] = 1
#         else:
#             combo[item] += 1
# print("form 1 element candidates "+str(combo))
# above_level_list = list()
# for key in combo:
#     if combo[key] >= level_of_support:
#         above_level_list.append(key)
# print("more then level of support "+str(above_level_list))
#
# n = 3
# for i in range(1,n):
#     if i > 1:
#         above_level_list = list(set([item for t in above_level_list for item in t]))
#
#     pairs = [",".join(map(str, comb)) for comb in combinations(above_level_list, i+1)]
#     print("form pairs "+str(pairs))
#
#     itemset = list()
#     for pair in pairs:
#         itemset.append(pair.split(","))
#
#     combo.clear()
#     above_level_list.clear()
#     for key in sets:
#         for item in itemset:
#             if all(elem in sets[key].split(",") for elem in item):
#                 if tuple(item) not in combo:
#                     combo[tuple(item)] = 1
#                 else:
#                     combo[tuple(item)] += 1
#
#     print("level of support of each combination "+str(combo))
#
#     for key in combo:
#         if combo[key] >= level_of_support:
#             above_level_list.append(key)
#     print("more then level of support "+str(above_level_list))

class Individual:
    def __init__(self, products: List[str]):
        self.products = products
        self.fitness = 0

    def __lt__(self, other):
        return self.fitness < other.fitness

    def __gt__(self, other):
        return self.fitness > other.fitness


def compute_fitness_values(individual, all_products):
    fitness = 0
    for products in all_products:
        product_count = 0
        for individual_product in individual.products:
            if individual_product in products:
                product_count += 1
        if product_count == len(individual.products):
            fitness += 1
    individual.fitness = fitness


def tournament_fight(inner_population):
    first = int(random() * len(inner_population))
    second = int(random() * len(inner_population))
    third = int(random() * len(inner_population))
    fourth = int(random() * len(inner_population))
    selected = [first]
    while second in selected:
        second = int(random() * len(inner_population))
    selected.append(second)
    while third in selected:
        third = int(random() * len(inner_population))
    selected.append(third)
    while fourth == first:
        fourth = int(random() * len(inner_population))

    individual1 = inner_population[first]
    individual2 = population[second]
    individual3 = population[third]
    individual4 = population[fourth]

    result = [individual1 if individual1 > individual2 else individual2,
              individual3 if individual3 > individual4 else individual4]
    return result


def create_new_individuals(parent1: Individual, parent2: Individual):
    first = Individual([*parent1.products[0:2], *parent2.products[2:]])
    second = Individual([*parent1.products[2:], *parent2.products[:2]])
    return first, second


def add_to_population(parents, children, inner_population):
    for child in children:
        for parent in parents:
            if child > parent:
                if parent in inner_population:
                    inner_population.remove(parent)
                inner_population.append(child)


def mutate(inner_population, all_products):
    for ind in inner_population:
        if random() < 0.01:
            while True:
                random_products_index = int(random() * len(all_products))
                random_product_index = int(random() * len(all_products[random_products_index]))
                random_product = all_products[random_products_index][random_product_index]
                if random_product not in ind.products:
                    random_genome = int(random() * len(ind.products))
                    ind.products[random_genome] = random_product
                else:
                    continue
                break


start = time()
workbook = load_workbook('Online Retail.xlsx')
sheet = workbook.active
sets = dict()
number_of_generations = 4

for rowno, rowval in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
    invoice_number = rowval[0].value
    if str(invoice_number).startswith("C"):
        continue
    stock_code = rowval[1].value
    customer_id = rowval[6].value
    if customer_id is None:
        continue
    if customer_id in sets:
        sets[customer_id].append(stock_code)
    else:
        sets[customer_id] = [stock_code]

population = list()
for key, value in sets.items():
    if len(value) == 3:
        population.append(Individual(value))

for individual in population:
    compute_fitness_values(individual, list(sets.values()))

for generation in range(number_of_generations):
    for k in range(len(population)):
        parent1, parent2 = tournament_fight(population)
        new_individual1, new_individual2 = create_new_individuals(parent1, parent2)
        compute_fitness_values(new_individual1, list(sets.values()))
        compute_fitness_values(new_individual2, list(sets.values()))
        add_to_population([parent1, parent2], [new_individual1, new_individual2], population)
        mutate(population, list(sets.values()))
print("Time: " + str(time() - start))
for person in population:
    print(person.products)
